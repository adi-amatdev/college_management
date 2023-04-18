from django.contrib import messages
from django.contrib.auth import authenticate,login,logout
from django.http import HttpResponse
from django.shortcuts import render
from django.shortcuts import HttpResponseRedirect
from student_management_app.EmailBackEnd import EmailBackEnd
from django.shortcuts import render, redirect
from django.http import HttpResponseRedirect
from openpyxl import load_workbook
from django.db import transaction
# from .models import TestScores

from rest_framework.decorators import api_view
from rest_framework import status 
from rest_framework.response import Response

import requests

from .serializers import *
from .models import *
from django.http import JsonResponse

def showdemopage(request):
    return render(request,"demo.html")

def showloginpage(request):
    return render(request,"main_login.html")

def doLogin(request):
    if request.method != 'POST':
        return HttpResponse("<h2>Method not allowed</h2>")
    else:
        user = EmailBackEnd.authenticate(request,username=request.POST.get("email"),password = request.POST.get("password"))
        if user!=None:
            login(request,user)
            if user.user_type=="1":
                return HttpResponseRedirect('/admin_home')
            elif user.user_type=="2":
                return HttpResponseRedirect('/staff_home')
            else:
                return HttpResponseRedirect('/student_home')
        else:
            messages.error(request,"Invalid Login Credentials")
            return HttpResponseRedirect("/")
        
def getuserdetails(request):
    if request.user!=None:
        return HttpResponse("User : "+request.user.email + " usertype : " + request.user.user_type)
    else:
        return HttpResponse("Please Login First")
    
def logout_user(request):
    logout(request)
    return HttpResponseRedirect("/")

# @transaction.atomic
# def upload_scores(request):
#     if request.method == 'POST' and request.FILES['scores_file']:
#         # Get the uploaded file from the request
#         scores_file = request.FILES['scores_file']
#         # Load the workbook using openpyxl
#         wb = load_workbook(scores_file)
#         # Get the active worksheet
#         ws = wb.active
#         # Loop through the rows in the worksheet, starting from the second row
#         for row in ws.iter_rows(min_row=2, values_only=True):
#             # Create a new TestScores object and save it to the database
#             test_scores = TestScores(test1=row[0], test2=row[1], test3=row[2])
#             test_scores.save()
#     # Redirect back to the upload page
#         return HttpResponseRedirect('/upload/')
#     else:
#         return render(request, 'upload.html')